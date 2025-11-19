import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import shutil
import numpy as np

# ======================
# CONFIG
# ======================

TEMPLATE_FILE = "Systeme_Entrainement_Empereur_ULTIME.xlsx"
DATA_FILE = "empereur_data.xlsx"

# ======================
# EXERCICES (doivent matcher l'Excel V3)
# ======================

LEGS_EXOS = [
    "Front Squat (wedge)",
    "Back Squat",
    "Snatch Grip Deadlift (position haute)",
    "Bulgarian Split Squat haltères",
    "Hack Squat",
    "Leg Press",
    "Leg Extension (full stretch)",
    "Leg Curl allongé",
    "Leg Curl assis",
    "Mollets debout",
    "Mollets assis",
    "Belt Squat",
    "Romanian Deadlift (barre)",
    "Hip Thrust barre",
    "Cable Kickback",
    "Abduction machine",
    "Standing Hip Abduction",
]

PUSH_EXOS = [
    "Développé couché barre / haltères",
    "Développé militaire barre / haltères",
    "Développé incliné batte / haltères",
    "Développé Arnold",
    "Kickbacks triceps",
    "Pompes",
    "Pompes lestées",
    "Pompes diamants",
    "Dips",
    "Dips lestées",
    "Chest-to-wall Hold",
    "Handstand Hold",
    "Pike push-up",
    "HSPU Négative",
    "HSPU partiels (mur)",
    "HSPU",
    "HSPU lestés",
    "Écarté incliné à la poulie",
    "Élévations latérales",
    "Extension triceps poulie",
]

PULL_EXOS = [
    "Tractions",
    "Tractions lestées",
    "Muscle-up",
    "Muscles-up lestées",
    "Rowing barre pronation",
    "Rowing machine unilatérale",
    "Good Morning barre basse",
    "Tirage vertical poulie inversée",
    "Curl biceps haltères",
    "Curl marteau haltères",
    "Face Pulls",
    "Shrugs lourds",
    "OMAD",  # Oiseau machine arrière d’épaules
]

FULL_EXOS = [
    "Box Jump",
    "Tuck Jumps",
    "Pistol Squat D",
    "Pistol Squat G",
    "Step-up genou haut D",
    "Step-up genou haut G",
    "High knees explosifs D",
    "High knees explosifs G",
    "Farmer Walk lourd",
    "Burpees",
    "Développé militaire au poids du corps",
    "Dips coréen",
    "Pompes inclinées pieds surélevés",
]

# Modes pour les pages de séance (kg/reps/sec)
LEGS_MODES = {ex: "kg_reps" for ex in LEGS_EXOS}

PUSH_MODES = {ex: "kg_reps" for ex in PUSH_EXOS}
for ex in ["Pompes", "Pompes diamants", "Dips", "Pike push-up",
           "HSPU Négative", "HSPU partiels (mur)", "HSPU"]:
    PUSH_MODES[ex] = "reps_only"
PUSH_MODES["Chest-to-wall Hold"] = "sec_only"
PUSH_MODES["Handstand Hold"] = "sec_only"

PULL_MODES = {ex: "kg_reps" for ex in PULL_EXOS}
for ex in ["Tractions", "Muscle-up"]:
    PULL_MODES[ex] = "reps_only"

FULL_MODES = {ex: "kg_reps" for ex in FULL_EXOS}
FULL_MODES["Farmer Walk lourd"] = "kg_only"


# ======================
# FICHIERS
# ======================

def get_excel_file(data_only=False):
    """Utilise une copie DATA_FILE modifiable.
    Si absente, on la crée à partir du TEMPLATE_FILE.
    """
    template_path = Path(TEMPLATE_FILE)
    if not template_path.exists():
        st.error(f"Fichier modèle introuvable : {template_path.resolve()}")
        st.stop()

    data_path = Path(DATA_FILE)
    if not data_path.exists():
        shutil.copy(template_path, data_path)

    wb = load_workbook(data_path, data_only=data_only)
    return wb, data_path


# ======================
# UTILITAIRES
# ======================

def get_next_lifestyle_day(ws):
    """Retourne le prochain jour à utiliser en ne tenant compte
    que des lignes où il y a vraiment des données Lifestyle.
    On ignore les formules de Readiness en colonne 9.
    """
    last = 0
    for row in range(2, ws.max_row + 1):
        jour = ws.cell(row=row, column=1).value
        if not isinstance(jour, int):
            continue

        # On regarde s'il y a au moins une donnée réelle (Sommeil à Humeur, colonnes 2 à 8)
        has_data = False
        for col in range(2, 9):  # on ignore la colonne 9 qui contient une formule par défaut
            if ws.cell(row=row, column=col).value not in (None, ""):
                has_data = True
                break

        if has_data and jour > last:
            last = jour

    # Si rien de rempli -> on commence à 1
    return last + 1 if last > 0 else 1


def _to_float(series):
    return pd.to_numeric(series, errors="coerce")


def epley(kg, reps):
    return kg * (1 + reps / 30.0)


# ======================
# LIFESTYLE
# ======================

def page_lifestyle():
    st.header("📋 Lifestyle – Saisie quotidienne")

    wb, data_path = get_excel_file()
    ws = wb["Lifestyle"]

    jour = get_next_lifestyle_day(ws)
    st.info(f"Jour enregistré : **{jour}** (prochain enregistrement)")

    col1, col2 = st.columns(2)
    with col1:
        sommeil = st.number_input("Sommeil (0-10)", 0.0, 10.0, 7.0, 0.5)
        hydrat = st.number_input("Hydratation (0-10)", 0.0, 10.0, 8.0, 0.5)
        nutri = st.number_input("Nutrition (0-10)", 0.0, 10.0, 7.0, 0.5)
        stress = st.number_input("Stress (0-10, plus = pire)", 0.0, 10.0, 3.0, 0.5)
    with col2:
        conc = st.number_input("Concentration (0-10)", 0.0, 10.0, 7.0, 0.5)
        energie = st.number_input("Énergie (0-10)", 0.0, 10.0, 7.0, 0.5)
        humeur = st.number_input("Humeur (0-10)", 0.0, 10.0, 7.0, 0.5)

    if st.button("💾 Enregistrer Lifestyle"):
        row = None
        for r in range(2, ws.max_row + 2):
            if ws.cell(row=r, column=1).value is None:
                row = r
                break
        if row is None:
            row = ws.max_row + 1

        s = float(sommeil)
        h = float(hydrat)
        n = float(nutri)
        stv = float(stress)
        c = float(conc)
        e = float(energie)
        hm = float(humeur)

        ws.cell(row=row, column=1).value = jour
        ws.cell(row=row, column=2).value = s
        ws.cell(row=row, column=3).value = h
        ws.cell(row=row, column=4).value = n
        ws.cell(row=row, column=5).value = stv
        ws.cell(row=row, column=6).value = c
        ws.cell(row=row, column=7).value = e
        ws.cell(row=row, column=8).value = hm

        if ws.cell(row=1, column=9).value in (None, ""):
            ws.cell(row=1, column=9).value = "Readiness"

        score_pos = (s + h + n + c + e + hm) / 6.0
        score_stress = 10.0 - stv
        readiness10 = 0.7 * score_pos + 0.3 * score_stress
        readiness100 = round(readiness10 * 10)

        ws.cell(row=row, column=9).value = readiness100

        wb.save(data_path)
        st.success(f"Lifestyle jour {jour} enregistré. Readiness = {readiness100}/100")


# ======================
# RPE EXAM & DB
# ======================

def rpe_from_max(val, unit):
    """Retourne un dict {5: v5, ..., 10: v10} à partir d'une valeur max.
    Pour les kg : pourcentages plus fins.
    Pour reps/sec : proportion linéaire.
    """
    if val is None:
        return {r: None for r in range(5, 11)}

    if unit == "kg":
        factors = {
            5: 0.80,
            6: 0.86,
            7: 0.90,
            8: 0.94,
            9: 0.97,
            10: 1.00,
        }
        return {r: round(val * factors[r], 1) for r in range(5, 11)}
    else:  # reps ou sec
        factors = {
            5: 0.50,
            6: 0.60,
            7: 0.70,
            8: 0.80,
            9: 0.90,
            10: 1.00,
        }
        return {r: int(round(val * factors[r])) for r in range(5, 11)}


def recompute_rpe_database(wb, data_path):
    """Lit RPE_EXAM, applique la logique de calcul + propagation HSPU,
    et écrit RPE_DATABASE.
    """
    df_exam = pd.read_excel(data_path, sheet_name="RPE_EXAM")

    max_map = {}

    for _, row in df_exam.iterrows():
        ex = row["Exercice"]
        unit = row["Unit"]
        max_kg = row.get("Max_kg")
        max_reps = row.get("Max_reps")
        max_sec = row.get("Max_sec")

        val = None
        if unit == "kg" and pd.notna(max_kg):
            val = float(max_kg)
        elif unit == "reps" and pd.notna(max_reps):
            val = float(max_reps)
        elif unit == "sec" and pd.notna(max_sec):
            val = float(max_sec)

        max_map[ex] = (unit, val)

    pike = max_map.get("Pike push-up", (None, None))[1]
    hspu_neg_unit, hspu_neg_val = max_map.get("HSPU Négative", ("reps", None))
    hspu_part_unit, hspu_part_val = max_map.get("HSPU partiels (mur)", ("reps", None))
    hspu_unit, hspu_val = max_map.get("HSPU", ("reps", None))

    if pike is not None and hspu_neg_val is None:
        hspu_neg_val = max(1, int(round(pike / 3)))
        max_map["HSPU Négative"] = ("reps", hspu_neg_val)

    if hspu_neg_val is not None and hspu_part_val is None:
        hspu_part_val = max(1, int(round(hspu_neg_val / 2)))
        max_map["HSPU partiels (mur)"] = ("reps", hspu_part_val)

    if hspu_part_val is not None and hspu_val is None:
        hspu_val = max(1, int(round(hspu_part_val / 2)))
        max_map["HSPU"] = ("reps", hspu_val)

    rows = []
    for _, row in df_exam.iterrows():
        ex = row["Exercice"]
        cat = row["Category"]
        unit, base_val = max_map.get(ex, (row["Unit"], None))
        rpes = rpe_from_max(base_val, unit)
        rows.append({
            "Exercice": ex,
            "Category": cat,
            "Unit": unit,
            "RPE5": rpes[5],
            "RPE6": rpes[6],
            "RPE7": rpes[7],
            "RPE8": rpes[8],
            "RPE9": rpes[9],
            "RPE10": rpes[10],
        })

    df_db = pd.DataFrame(rows, columns=["Exercice", "Category", "Unit",
                                        "RPE5", "RPE6", "RPE7", "RPE8", "RPE9", "RPE10"])

    if "RPE_DATABASE" not in wb.sheetnames:
        ws_db = wb.create_sheet("RPE_DATABASE")
    else:
        ws_db = wb["RPE_DATABASE"]
        ws_db.delete_rows(1, ws_db.max_row)

    ws_db.append(["Exercice", "Category", "Unit", "RPE5", "RPE6", "RPE7", "RPE8", "RPE9", "RPE10"])
    for _, r in df_db.iterrows():
        ws_db.append(list(r.values))

    wb.save(data_path)


def page_rpe_exam():
    st.header("🎯 RPE EXAM – Tests de référence")

    wb, data_path = get_excel_file()

    st.markdown("**Entre uniquement les exos que tu as testés.** Les autres resteront avec leurs anciennes valeurs.")

    def bloc_exam(title, exos, rules_key_prefix):
        st.subheader(title)
        for ex in exos:
            cols = st.columns(3)
            cols[0].markdown(f"**{ex}**")
            kg_field, reps_field, sec_field = None, None, None

            if title.startswith("EXAMENS LEGS"):
                kg_field = cols[1].text_input("kg", key=f"{rules_key_prefix}_{ex}_kg")
                reps_field = cols[2].text_input("reps", key=f"{rules_key_prefix}_{ex}_reps")
            elif title.startswith("EXAMENS FULL"):
                if ex == "Farmer Walk lourd":
                    kg_field = cols[1].text_input("kg", key=f"{rules_key_prefix}_{ex}_kg")
                else:
                    kg_field = cols[1].text_input("kg", key=f"{rules_key_prefix}_{ex}_kg")
                    reps_field = cols[2].text_input("reps", key=f"{rules_key_prefix}_{ex}_reps")
            elif title.startswith("EXAMENS PUSH"):
                if ex in ["Pompes", "Pompes diamants", "Dips"]:
                    reps_field = cols[2].text_input("reps", key=f"{rules_key_prefix}_{ex}_reps")
                elif ex in ["Chest-to-wall Hold", "Handstand Hold"]:
                    sec_field = cols[2].text_input("sec", key=f"{rules_key_prefix}_{ex}_sec")
                elif ex in ["Pike push-up", "HSPU Négative", "HSPU partiels (mur)", "HSPU"]:
                    reps_field = cols[2].text_input("reps", key=f"{rules_key_prefix}_{ex}_reps")
                else:
                    kg_field = cols[1].text_input("kg", key=f"{rules_key_prefix}_{ex}_kg")
                    reps_field = cols[2].text_input("reps", key=f"{rules_key_prefix}_{ex}_reps")
            elif title.startswith("EXAMENS PULL"):
                if ex in ["Tractions", "Muscle-up"]:
                    reps_field = cols[2].text_input("reps", key=f"{rules_key_prefix}_{ex}_reps")
                else:
                    kg_field = cols[1].text_input("kg", key=f"{rules_key_prefix}_{ex}_kg")
                    reps_field = cols[2].text_input("reps", key=f"{rules_key_prefix}_{ex}_reps")

    bloc_exam("EXAMENS LEGS RPE :", LEGS_EXOS, "LEGS")
    st.markdown("---")
    bloc_exam("EXAMENS PUSH RPE :", PUSH_EXOS, "PUSH")
    st.markdown("---")
    bloc_exam("EXAMENS PULL RPE :", PULL_EXOS, "PULL")
    st.markdown("---")
    bloc_exam("EXAMENS FULL RPE :", FULL_EXOS, "FULL")

    if st.button("✅ Valider les examens RPE"):
        ws = wb["RPE_EXAM"]

        ex_row_map = {}
        for r in range(2, ws.max_row + 1):
            ex_name = ws.cell(row=r, column=1).value
            ex_row_map[ex_name] = r

        def update_exos(exos, prefix):
            for ex in exos:
                row = ex_row_map.get(ex)
                if not row:
                    continue
                kg_key = f"{prefix}_{ex}_kg"
                reps_key = f"{prefix}_{ex}_reps"
                sec_key = f"{prefix}_{ex}_sec"
                kg_val = st.session_state.get(kg_key, "").strip() if kg_key in st.session_state else ""
                reps_val = st.session_state.get(reps_key, "").strip() if reps_key in st.session_state else ""
                sec_val = st.session_state.get(sec_key, "").strip() if sec_key in st.session_state else ""

                if kg_val != "":
                    try:
                        ws.cell(row=row, column=3).value = float(kg_val)
                    except ValueError:
                        pass
                if reps_val != "":
                    try:
                        ws.cell(row=row, column=4).value = float(reps_val)
                    except ValueError:
                        pass
                if sec_val != "":
                    try:
                        ws.cell(row=row, column=5).value = float(sec_val)
                    except ValueError:
                        pass

        update_exos(LEGS_EXOS, "LEGS")
        update_exos(PUSH_EXOS, "PUSH")
        update_exos(PULL_EXOS, "PULL")
        update_exos(FULL_EXOS, "FULL")

        wb.save(data_path)
        recompute_rpe_database(wb, data_path)
        st.success("Examens RPE mis à jour et base de données RPE recalculée.")


def page_rpe_database():
    st.header("📚 BASE DE DONNÉE – RPE 5 à 10")

    wb, data_path = get_excel_file(data_only=True)

    try:
        df_db = pd.read_excel(data_path, sheet_name="RPE_DATABASE")
    except Exception as e:
        st.warning(f"Impossible de lire RPE_DATABASE : {e}")
        return

    def show_block(title, category):
        st.subheader(title)
        sub = df_db[df_db["Category"] == category].copy()
        if sub.empty:
            st.info("Aucune donnée pour l'instant.")
            return
        sub = sub[["Exercice", "Unit", "RPE5", "RPE6", "RPE7", "RPE8", "RPE9", "RPE10"]]
        sub = sub.rename(columns={
            "Exercice": "Nom de l’exercice",
            "Unit": "Unité"
        })
        st.dataframe(sub, use_container_width=True)

    show_block("RÉSULTATS LEGS RPE", "LEGS")
    st.markdown("---")
    show_block("RÉSULTATS PUSH RPE", "PUSH")
    st.markdown("---")
    show_block("RÉSULTATS PULL RPE", "PULL")
    st.markdown("---")
    show_block("RÉSULTATS FULL RPE", "FULL")


# ======================
# PAGES SEANCES
# ======================

def find_or_create_session_row(ws, session_number: int):
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == session_number:
            return r
    new_row = ws.max_row + 1 if ws.max_row >= 2 else 2
    ws.cell(row=new_row, column=1).value = session_number
    return new_row


def page_seance_generic(title, sheet_name, exos, modes):
    st.header(title)

    wb, data_path = get_excel_file()
    if sheet_name not in wb.sheetnames:
        st.error(f"Feuille '{sheet_name}' introuvable dans Excel.")
        return
    ws = wb[sheet_name]

    session = st.number_input("Numéro de séance", min_value=1, step=1, value=1)
    st.write("Remplis uniquement les exercices faits. Laisse vide pour ignorer.")

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    inputs = []

    for ex in exos:
        mode = modes.get(ex, "kg_reps")
        cols = st.columns(3)
        cols[0].markdown(f"**{ex}**")
        if mode in ("kg_reps", "kg_only"):
            kg_col = f"{ex} (kg)"
            kg_str = cols[1].text_input("kg", key=f"{sheet_name}_{session}_{ex}_kg")
            inputs.append((kg_col, "kg", kg_str))
        if mode in ("kg_reps", "reps_only"):
            reps_col = f"{ex} (reps)"
            reps_str = cols[2].text_input("reps", key=f"{sheet_name}_{session}_{ex}_reps")
            inputs.append((reps_col, "reps", reps_str))
        if mode == "sec_only":
            sec_col = f"{ex} (sec)"
            sec_str = cols[2].text_input("sec", key=f"{sheet_name}_{session}_{ex}_sec")
            inputs.append((sec_col, "sec", sec_str))

    if st.button(f"💾 Enregistrer {title}"):
        row = find_or_create_session_row(ws, int(session))

        for col_name, vtype, sval in inputs:
            sval = sval.strip()
            if sval == "":
                continue
            col_idx = headers.get(col_name)
            if not col_idx:
                continue
            try:
                if vtype == "kg":
                    val = float(sval)
                else:
                    val = int(float(sval))
                ws.cell(row=row, column=col_idx).value = val
            except ValueError:
                continue

        wb.save(data_path)
        st.success(f"{title} – Séance {int(session)} enregistrée.")


def page_seance_legs():
    page_seance_generic("SÉANCE LEGS", "Seance_Legs", LEGS_EXOS, LEGS_MODES)


def page_seance_push():
    page_seance_generic("SÉANCE PUSH", "Seance_Push", PUSH_EXOS, PUSH_MODES)


def page_seance_pull():
    page_seance_generic("SÉANCE PULL", "Seance_Pull", PULL_EXOS, PULL_MODES)


def page_seance_full():
    page_seance_generic("SÉANCE FULL", "Seance_Full", FULL_EXOS, FULL_MODES)


# ======================
# METRIQUES : CHARGE, FATIGUE, SAH V2
# ======================

def load_all_sessions_wide(data_path: Path):
    frames = []
    for sheet in ["Seance_Legs", "Seance_Push", "Seance_Pull", "Seance_Full"]:
        try:
            df = pd.read_excel(data_path, sheet_name=sheet)
            df["Séance"] = pd.to_numeric(df["Séance"], errors="coerce")
            df = df.dropna(subset=["Séance"])
            frames.append(df)
        except Exception:
            continue
    if not frames:
        return None
    df_all = pd.concat(frames, ignore_index=True)
    df_all["Séance"] = df_all["Séance"].astype(int)
    df_all = df_all.sort_values("Séance")
    return df_all


def compute_session_metrics(data_path: Path):
    df_all = load_all_sessions_wide(data_path)
    if df_all is None:
        return None

    loads = {}

    for _, row in df_all.iterrows():
        s = int(row["Séance"])
        load_row = 0.0
        for col in df_all.columns:
            if col == "Séance":
                continue
            val = row[col]
            if pd.isna(val):
                continue

            if col.endswith(" (kg)"):
                base = float(val)
                base_name = col[:-5]
                reps_col = base_name + " (reps)"
                if reps_col in df_all.columns:
                    reps = row.get(reps_col)
                    if pd.notna(reps):
                        load_row += base * float(reps)
                    else:
                        load_row += base
                else:
                    load_row += base
            elif col.endswith(" (reps)") or col.endswith(" (sec)"):
                load_row += float(val)

        loads[s] = loads.get(s, 0.0) + load_row

    if not loads:
        return None

    df_sessions = pd.DataFrame(
        [{"Séance": k, "Load": v} for k, v in loads.items()]
    ).sort_values("Séance")

    return df_sessions


def compute_fatigue_metrics(data_path: Path, window: int = 7):
    df_s = compute_session_metrics(data_path)
    if df_s is None or df_s.empty:
        return None, None, None

    loads = df_s["Load"].values
    if len(loads) >= window:
        loads_window = loads[-window:]
    else:
        loads_window = loads

    mean_load = float(np.mean(loads_window))
    std_load = float(np.std(loads_window)) if len(loads_window) > 1 else 0.0

    if std_load == 0:
        monotony = 0.0
    else:
        monotony = mean_load / std_load

    strain = mean_load * monotony
    return mean_load, monotony, strain


def safe_nanmax(arr):
    arr = np.array(arr, dtype=float)
    if arr.size == 0 or np.isnan(arr).all():
        return 0.0
    return float(np.nanmax(arr))


def compute_sah_v2(data_path: Path):
    df_all = load_all_sessions_wide(data_path)
    if df_all is None:
        return None, {}

    sq_back_kg = _to_float(df_all.get("Back Squat (kg)"))
    sq_back_reps = _to_float(df_all.get("Back Squat (reps)"))
    sq_front_kg = _to_float(df_all.get("Front Squat (wedge) (kg)"))
    sq_front_reps = _to_float(df_all.get("Front Squat (wedge) (reps)"))
    bench_kg = _to_float(df_all.get("Développé couché barre / haltères (kg)"))
    bench_reps = _to_float(df_all.get("Développé couché barre / haltères (reps)"))
    roman_kg = _to_float(df_all.get("Romanian Deadlift (barre) (kg)"))
    roman_reps = _to_float(df_all.get("Romanian Deadlift (barre) (reps)"))

    squat1_back = epley(sq_back_kg, sq_back_reps)
    squat1_front = epley(sq_front_kg, sq_front_reps)
    if squat1_back.size and squat1_front.size:
        squat_1rm = np.nanmax(np.vstack([squat1_back, squat1_front]), axis=0)
    else:
        squat_1rm = squat1_back if squat1_back.size else squat1_front

    bench_1rm = epley(bench_kg, bench_reps)
    dead_1rm = epley(roman_kg, roman_reps)

    best_squat = safe_nanmax(squat_1rm)
    best_bench = safe_nanmax(bench_1rm)
    best_dead = safe_nanmax(dead_1rm)

    sq_target = 220.0
    bp_target = 160.0
    dl_target = 260.0

    str_squat = min(best_squat / sq_target, 1.3) if sq_target > 0 else 0
    str_bench = min(best_bench / bp_target, 1.3) if bp_target > 0 else 0
    str_dead = min(best_dead / dl_target, 1.3) if dl_target > 0 else 0

    strength_index = float(np.mean([str_squat, str_bench, str_dead]) * 100.0)

    details = {
        "Squat1RM": round(best_squat, 1),
        "Bench1RM": round(best_bench, 1),
        "Dead1RM": round(best_dead, 1),
        "StrengthIndex": round(strength_index, 1),
    }

    hspu_reps = _to_float(df_all.get("HSPU (reps)"))
    mu_reps = _to_float(df_all.get("Muscle-up (reps)"))
    t_lest_kg = _to_float(df_all.get("Tractions lestées (kg)"))

    best_hspu = safe_nanmax(hspu_reps)
    best_mu = safe_nanmax(mu_reps)
    best_tlest = safe_nanmax(t_lest_kg)

    details.update({
        "HSPU": best_hspu,
        "MU": best_mu,
        "TractionLestee": best_tlest,
    })

    hspu_target = 20.0
    mu_target = 10.0
    tlest_target = 80.0

    s_hspu = min(best_hspu / hspu_target, 1.3) if hspu_target > 0 else 0
    s_mu = min(best_mu / mu_target, 1.3) if mu_target > 0 else 0
    s_tlest = min(best_tlest / tlest_target, 1.3) if tlest_target > 0 else 0

    skill_index = float(np.mean([s_hspu, s_mu, s_tlest]) * 100.0)
    power_index = float(np.mean([s_mu, s_tlest]) * 100.0)

    details["SkillIndex"] = round(skill_index, 1)
    details["PowerIndex"] = round(power_index, 1)

    sah_components = [strength_index, skill_index, power_index]
    weights = [0.4, 0.4, 0.2]
    sah_v2 = float(np.average(sah_components, weights=weights))
    sah_v2 = float(np.clip(sah_v2, 0, 100))
    details["SAH_V2"] = round(sah_v2, 1)

    return sah_v2, details


def classify_skill_level(skill_index: float):
    if skill_index is None:
        return "Inconnu"
    if skill_index < 30:
        return "Débutant"
    if skill_index < 60:
        return "Intermédiaire"
    if skill_index < 85:
        return "Avancé"
    return "Élite"


def get_latest_readiness(data_path: Path):
    try:
        df_life = pd.read_excel(data_path, sheet_name="Lifestyle")
    except Exception:
        return None
    col = None
    if "Readiness" in df_life.columns:
        col = "Readiness"
    elif df_life.shape[1] >= 9:
        col = df_life.columns[8]
    if col is None:
        return None
    vals = pd.to_numeric(df_life[col], errors="coerce").dropna()
    if vals.empty:
        return None
    return float(vals.iloc[-1])


def get_last_session_info(data_path: Path):
    df_s = compute_session_metrics(data_path)
    if df_s is None or df_s.empty:
        return None
    last = df_s.iloc[-1]
    return {
        "Séance": int(last["Séance"]),
        "Load": float(last["Load"]),
    }


# ======================
# DASHBOARDS
# ======================

def page_dashboards():
    st.header("📊 Dashboards – Volume, 1RM, Calisthénie")

    wb, data_path = get_excel_file(data_only=True)

    df_s = compute_session_metrics(data_path)
    if df_s is None or df_s.empty:
        st.info("Aucune séance enregistrée pour l'instant.")
        return

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Volume par séance (Load total)")
        st.line_chart(df_s.set_index("Séance")["Load"])

    df_all = load_all_sessions_wide(data_path)
    if df_all is None:
        return

    sq_back_kg = _to_float(df_all.get("Back Squat (kg)"))
    sq_back_reps = _to_float(df_all.get("Back Squat (reps)"))
    sq_front_kg = _to_float(df_all.get("Front Squat (wedge) (kg)"))
    sq_front_reps = _to_float(df_all.get("Front Squat (wedge) (reps)"))
    bench_kg = _to_float(df_all.get("Développé couché barre / haltères (kg)"))
    bench_reps = _to_float(df_all.get("Développé couché barre / haltères (reps)"))
    roman_kg = _to_float(df_all.get("Romanian Deadlift (barre) (kg)"))
    roman_reps = _to_float(df_all.get("Romanian Deadlift (barre) (reps)"))

    sq1_back = epley(sq_back_kg, sq_back_reps)
    sq1_front = epley(sq_front_kg, sq_front_reps)
    if sq1_back.size and sq1_front.size:
        sq1 = np.nanmax(np.vstack([sq1_back, sq1_front]), axis=0)
    else:
        sq1 = sq1_back if sq1_back.size else sq1_front

    bench1 = epley(bench_kg, bench_reps)
    dead1 = epley(roman_kg, roman_reps)

    df_1rm = pd.DataFrame({
        "Séance": df_all["Séance"],
        "Squat 1RM": sq1,
        "Bench 1RM": bench1,
        "Deadlift 1RM": dead1,
    }).dropna()

    with col2:
        st.subheader("1RM estimées (Epley)")
        if not df_1rm.empty:
            df_plot = df_1rm.groupby("Séance").max()[["Squat 1RM", "Bench 1RM", "Deadlift 1RM"]]
            st.line_chart(df_plot)
        else:
            st.info("Pas encore assez de données pour estimer les 1RM.")

    st.markdown("---")
    st.subheader("Indicateurs Calisthénie")
    hspu = _to_float(df_all.get("HSPU (reps)"))
    mu = _to_float(df_all.get("Muscle-up (reps)"))
    t_lest = _to_float(df_all.get("Tractions lestées (kg)"))

    df_cali = pd.DataFrame({
        "Séance": df_all["Séance"],
        "HSPU (reps)": hspu,
        "MU (reps)": mu,
        "Tractions lestées (kg)": t_lest,
    }).dropna(how="all", subset=["HSPU (reps)", "MU (reps)", "Tractions lestées (kg)"])

    if df_cali.empty:
        st.info("Pas encore de données calisthénie.")
    else:
        st.line_chart(df_cali.set_index("Séance"))


# ======================
# PR & SAH V2
# ======================

def page_pr_sah():
    st.header("🏆 PR & Score Athlète Hybride V2")

    wb, data_path = get_excel_file(data_only=True)

    sah_v2, details = compute_sah_v2(data_path)

    if sah_v2 is None:
        st.info("Pas encore assez de données (séances) pour calculer un SAH V2.")
        return

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("SAH V2", value=round(details.get("SAH_V2", sah_v2), 1))
    with col2:
        st.metric("StrengthIndex", value=details.get("StrengthIndex", "N/A"))
    with col3:
        st.metric("SkillIndex", value=details.get("SkillIndex", "N/A"))

    with st.expander("Détails complets SAH V2"):
        st.json(details)


# ======================
# PLANNING & SYNTHÈSE
# ======================

def page_planning():
    st.header("📅 Planning – Plan Annuel & Mésocycles")

    wb, data_path = get_excel_file(data_only=True)

    col1, col2 = st.columns(2)
    try:
        df_annuel = pd.read_excel(data_path, sheet_name="Plan Annuel")
        with col1:
            st.subheader("Plan Annuel")
            st.dataframe(df_annuel)
    except Exception as e:
        st.warning(f"Erreur lecture Plan Annuel : {e}")

    try:
        df_meso = pd.read_excel(data_path, sheet_name="Mésocycle-Type")
        with col2:
            st.subheader("Mésocycle-Type")
            st.dataframe(df_meso)
    except Exception as e:
        st.warning(f"Erreur lecture Mésocycle-Type : {e}")

    st.markdown("---")
    try:
        df_auto_meso = pd.read_excel(data_path, sheet_name="Auto-Mesocycles")
        st.subheader("Auto-Mesocycles")
        st.dataframe(df_auto_meso)
    except Exception as e:
        st.warning(f"Erreur lecture Auto-Mesocycles : {e}")


def page_reco_global():
    st.header("🧠 Synthèse & Recommandations globales")

    wb, data_path = get_excel_file(data_only=True)

    try:
        df_life = pd.read_excel(data_path, sheet_name="Lifestyle")
        if "Readiness" in df_life.columns:
            col = "Readiness"
        elif df_life.shape[1] >= 9:
            col = df_life.columns[8]
        else:
            col = None
        if col:
            vals = pd.to_numeric(df_life[col], errors="coerce").dropna()
            readiness_moy = float(vals.mean()) if not vals.empty else None
        else:
            readiness_moy = None
    except Exception:
        readiness_moy = None

    mean_load, monotony, strain = compute_fatigue_metrics(data_path)
    sah_v2, sah_details = compute_sah_v2(data_path)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Readiness moyen", value=round(readiness_moy, 1) if readiness_moy is not None else "N/A")
    with col2:
        st.metric("Charge moyenne (7 dernières séances)", value=int(mean_load) if mean_load is not None else "N/A")
    with col3:
        st.metric("Strain (7 dernières séances)", value=int(strain) if strain is not None else "N/A")

    col4, col5 = st.columns(2)
    with col4:
        st.metric("Monotony", value=round(monotony, 2) if monotony is not None else "N/A")
    with col5:
        st.metric("SAH V2", value=round(sah_details.get("SAH_V2", sah_v2), 1) if sah_v2 is not None else "N/A")

    st.markdown("---")
    st.subheader("Recommandation générale")

    if readiness_moy is None or mean_load is None or strain is None:
        st.info("Pas encore assez de données pour générer une recommandation complète.")
        return

    if readiness_moy >= 70 and strain < 20000:
        st.write("✅ Tu es dans une bonne zone pour pousser sur des séances lourdes ou de gros volume.")
    elif readiness_moy < 40 or strain >= 25000:
        st.write("⚠️ Zone de fatigue élevée : privilégie la gestion de la récupération, le skill propre ou le deload.")
    else:
        st.write("🟡 Zone intermédiaire : progression possible, mais surveille ton sommeil, stress et volumes.")


# ======================
# AUTO-SÉANCE INTELLIGENTE
# ======================

def compute_auto_seance_recommendation(data_path: Path, block_focus: str):
    readiness = get_latest_readiness(data_path)
    mean_load, monotony, strain = compute_fatigue_metrics(data_path)
    sah_v2, details = compute_sah_v2(data_path)
    last_info = get_last_session_info(data_path)

    skill_index = details.get("SkillIndex", 0.0)
    strength_index = details.get("StrengthIndex", 0.0)
    power_index = details.get("PowerIndex", 0.0)
    skill_level = classify_skill_level(skill_index)

    if readiness is None:
        readiness = 50.0
    if mean_load is None:
        mean_load = 0.0
    if monotony is None:
        monotony = 0.0
    if strain is None:
        strain = 0.0

    if readiness >= 70:
        readiness_zone = "High"
    elif readiness >= 40:
        readiness_zone = "Medium"
    else:
        readiness_zone = "Low"

    if strain >= 25000:
        strain_zone = "High"
    elif strain >= 10000:
        strain_zone = "Medium"
    else:
        strain_zone = "Low"

    if block_focus == "Force maximale":
        primary = "Force"
    elif block_focus == "Hypertrophie / Volume":
        primary = "Volume"
    elif block_focus == "Skill / Calisthénie":
        primary = "Skill"
    elif block_focus == "Puissance / Explosivité":
        primary = "Power"
    else:
        primary = "Deload"

    session_type = ""
    focus = ""
    intensity = ""
    volume_mod = ""
    rpe_target = ""
    notes = []
    structure = []

    if readiness_zone == "Low" or strain_zone == "High":
        if primary == "Deload":
            session_type = "Recovery / Off"
            focus = "Récupération globale"
            intensity = "Très basse"
            volume_mod = "20–40% du volume habituel"
            rpe_target = "RPE 5–6 max"
            notes.append("Fatigue ou strain élevés : privilégier la récupération active.")
            structure = [
                "20–30 min mobilité totale (hanches, épaules, colonne)",
                "10–20 min marche ou cardio très léger",
                "Travail technique très propre : handstand hold, supports, respiration",
                "Sauna / bain chaud / automassage si possible",
            ]
        else:
            session_type = "Skill / Recovery"
            focus = "Technique + Calisthénie propre + mobilité"
            intensity = "Basse à modérée"
            volume_mod = "40–60% du volume habituel"
            rpe_target = "RPE 6–7"
            notes.append("Readiness bas ou strain élevé : on garde la fréquence mais on baisse l'impact.")
            structure = [
                "Bloc skill : HSPU, MU, variations progressives",
                "Volume traction / push modéré, loin de l'échec",
                "Core & gainage (planche, hollow, arch)",
                "Long travail de stretching actif / PNF en fin de séance",
            ]
    else:
        if primary == "Force":
            session_type = "Heavy Strength"
            focus = "Force lourde (1–3 lifts principaux)"
            intensity = "Élevée"
            volume_mod = "70–90% du volume habituel"
            rpe_target = "RPE 8–9 sur les principaux mouvements"
            notes.append("Tu peux pousser lourd sur 1–3 exercices clés.")
            structure = [
                "1–2 mouvements principaux en 3–5 séries lourdes (3–6 reps)",
                "2–3 accessoires lourds ou modérés (6–10 reps)",
                "Un peu de skill en fin si énergie",
                "Mobilité / respiration pour redescendre le système",
            ]
        elif primary == "Volume":
            session_type = "Hypertrophie / Volume"
            focus = "Accumulation de volume contrôlé"
            intensity = "Modérée"
            volume_mod = "90–110% du volume habituel"
            rpe_target = "RPE 7–8"
            notes.append("Objectif : congestion et volume sans cramer le système nerveux.")
            structure = [
                "2 mouvements de base en 4×8–12",
                "3–4 exercices d'isolation (12–20 reps)",
                "Optionnel : finisher métabolique (farmer walk + burpees)",
                "Stretching ciblé sur les groupes très travaillés",
            ]
        elif primary == "Skill":
            session_type = "Skill Calisthénie"
            focus = "Maîtrise technique (HSPU / MU / équilibres)"
            intensity = "Modérée"
            volume_mod = "60–80% du volume habituel"
            rpe_target = "RPE 6–8, jamais à l'échec nerveux sur le skill"
            notes.append(f"Niveau skill actuel : {skill_level}. On consolide la technique.")
            structure = [
                "Bloc 1 : MU (progressions, 3–5 reps par série)",
                "Bloc 2 : HSPU / handstand (négatives, holds, partiels)",
                "Bloc 3 : tractions / dips / pompes pour volume contrôlé",
                "Mobility épaules + poignets en fin de séance",
            ]
        elif primary == "Power":
            session_type = "Puissance / Explosivité"
            focus = "Sauts, vitesse, intention explosive"
            intensity = "Élevée mais volume limité"
            volume_mod = "50–70% volume muscu, intensité maximale sur explosif"
            rpe_target = "RPE 7–8 (qualité, pas d'échec)"
            notes.append("Objectif : système nerveux rapide, pas cramé.")
            structure = [
                "Sauts (box jumps, broad jumps, 3–5 reps par série)",
                "Sprints courts / hill sprints si possible",
                "Un peu de force submax (70–80% 1RM, vitesse d'exécution)",
                "Mobilité hanches / chevilles",
            ]
        else:
            session_type = "Deload intelligent"
            focus = "Réduction de charge, maintien technique"
            intensity = "Basse à modérée"
            volume_mod = "40–60% du volume habituel"
            rpe_target = "RPE 6–7"
            notes.append("Bloc orienté gestion fatigue / décharge.")
            structure = [
                "Même structure qu'une séance normale mais -40% en charge/volume",
                "Travail technique plus propre (tempo, pauses)",
                "Beaucoup de mobilité / respiration en fin",
            ]

    if last_info is not None:
        notes.append(f"Dernière séance enregistrée : Séance {last_info['Séance']} – Load {int(last_info['Load'])}.")

    return {
        "readiness": readiness,
        "mean_load": mean_load,
        "monotony": monotony,
        "strain": strain,
        "sah_v2": sah_v2,
        "strength_index": strength_index,
        "skill_index": skill_index,
        "power_index": power_index,
        "skill_level": skill_level,
        "last_session": last_info,
        "session_type": session_type,
        "focus": focus,
        "intensity": intensity,
        "volume_mod": volume_mod,
        "rpe_target": rpe_target,
        "notes": notes,
        "structure_suggestion": structure,
    }


def page_auto_seance():
    st.header("🤖 Auto-Séance intelligente – Coach Empereur")

    wb, data_path = get_excel_file(data_only=True)

    st.markdown("Cette page te propose un **type de séance du jour** basé sur :")
    st.markdown("- Ta dernière valeur de **Readiness**")
    st.markdown("- La **charge** et le **strain** des 7 dernières séances")
    st.markdown("- Ton **niveau Skill** (calisthénie / puissance)")
    st.markdown("- L’**objectif du bloc** que tu choisis")

    block_focus = st.selectbox(
        "Objectif du bloc en cours",
        [
            "Force maximale",
            "Hypertrophie / Volume",
            "Skill / Calisthénie",
            "Puissance / Explosivité",
            "Déload / Gestion fatigue",
        ]
    )

    if st.button("⚡ Générer la séance recommandée"):
        reco = compute_auto_seance_recommendation(data_path, block_focus)

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Readiness (dernier jour)", value=round(reco["readiness"], 1))
        with col2:
            st.metric("Strain (7 dernières séances)", value=int(reco["strain"]))
        with col3:
            if reco["sah_v2"] is not None:
                st.metric("SAH V2", value=round(reco["sah_v2"], 1))
            else:
                st.metric("SAH V2", value="N/A")

        st.markdown("---")
        st.subheader("🧬 Profil actuel")
        col4, col5, col6 = st.columns(3)
        with col4:
            st.metric("StrengthIndex", value=round(reco["strength_index"], 1))
        with col5:
            st.metric("SkillIndex", value=round(reco["skill_index"], 1))
        with col6:
            st.metric("PowerIndex", value=round(reco["power_index"], 1))

        st.write(f"**Niveau Skill :** {reco['skill_level']}")

        if reco["last_session"] is not None:
            st.markdown("**Dernière séance enregistrée :**")
            st.json(reco["last_session"])

        st.markdown("---")
        st.subheader("📋 Séance du jour recommandée")

        st.write(f"**Type de séance :** {reco['session_type']}")
        st.write(f"**Focus :** {reco['focus']}")
        st.write(f"**Intensité :** {reco['intensity']}")
        st.write(f"**Volume relatif :** {reco['volume_mod']}")
        st.write(f"**RPE cible :** {reco['rpe_target']}")

        if reco["notes"]:
            st.markdown("**Notes du coach :**")
            for n in reco["notes"]:
                st.write(f"- {n}")

        if reco["structure_suggestion"]:
            st.markdown("**Structure suggérée :**")
            for s in reco["structure_suggestion"]:
                st.write(f"- {s}")


# ======================
# EXPORT & DEBUG
# ======================

def page_export_debug():
    st.header("📥 Export & Debug des données Empereur")

    wb, data_path = get_excel_file()

    st.subheader("Lifestyle – dernières entrées")
    try:
        df_life = pd.read_excel(data_path, sheet_name="Lifestyle")
        st.dataframe(df_life.tail(10))
    except Exception as e:
        st.warning(f"Impossible de lire Lifestyle : {e}")

    st.markdown("---")
    st.subheader("Séances LEGS – dernières entrées")
    try:
        df_legs = pd.read_excel(data_path, sheet_name="Seance_Legs")
        st.dataframe(df_legs.tail(10))
    except Exception as e:
        st.warning(f"Impossible de lire Seance_Legs : {e}")

    st.markdown("---")
    st.subheader("Séances PUSH – dernières entrées")
    try:
        df_push = pd.read_excel(data_path, sheet_name="Seance_Push")
        st.dataframe(df_push.tail(10))
    except Exception as e:
        st.warning(f"Impossible de lire Seance_Push : {e}")

    st.markdown("---")
    st.subheader("Séances PULL – dernières entrées")
    try:
        df_pull = pd.read_excel(data_path, sheet_name="Seance_Pull")
        st.dataframe(df_pull.tail(10))
    except Exception as e:
        st.warning(f"Impossible de lire Seance_Pull : {e}")

    st.markdown("---")
    st.subheader("Séances FULL – dernières entrées")
    try:
        df_full = pd.read_excel(data_path, sheet_name="Seance_Full")
        st.dataframe(df_full.tail(10))
    except Exception as e:
        st.warning(f"Impossible de lire Seance_Full : {e}")

    st.markdown("---")
    st.subheader("RPE_EXAM & RPE_DATABASE – aperçu")
    try:
        df_exam = pd.read_excel(data_path, sheet_name="RPE_EXAM")
        st.write("RPE_EXAM")
        st.dataframe(df_exam.head(20))
    except Exception as e:
        st.warning(f"Impossible de lire RPE_EXAM : {e}")

    try:
        df_db = pd.read_excel(data_path, sheet_name="RPE_DATABASE")
        st.write("RPE_DATABASE")
        st.dataframe(df_db.head(20))
    except Exception as e:
        st.warning(f"Impossible de lire RPE_DATABASE : {e}")

    st.markdown("---")
    st.subheader("Télécharger le fichier de données complet")

    data_path = Path(DATA_FILE)
    if not data_path.exists():
        st.info("Aucun fichier empereur_data.xlsx trouvé pour l'instant (enregistre d'abord des données).")
    else:
        with open(data_path, "rb") as f:
            binary = f.read()
        st.download_button(
            label="📥 Télécharger empereur_data.xlsx",
            data=binary,
            file_name="empereur_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.markdown("---")
    st.subheader("♻️ Réinitialiser toutes les données")

    st.warning(
        "⚠️ Cette action supprime **toutes** les données actuelles (Lifestyle, Séances, RPE, etc.) "
        "et recrée un fichier vierge à partir du modèle."
    )

    if st.button("🔴 Réinitialiser empereur_data.xlsx"):
        if data_path.exists():
            data_path.unlink()
            st.success(
                "Toutes les données ont été réinitialisées. "
                "La prochaine utilisation de l'app recréera un fichier vierge à partir du modèle."
            )
        else:
            st.info("Aucun fichier de données à supprimer.")


# ======================
# MAIN
# ======================

PAGES = {
    "Lifestyle": page_lifestyle,
    "RPE EXAM": page_rpe_exam,
    "BASE DE DONNÉE": page_rpe_database,
    "SÉANCE LEGS": page_seance_legs,
    "SÉANCE PUSH": page_seance_push,
    "SÉANCE PULL": page_seance_pull,
    "SÉANCE FULL": page_seance_full,
    "Dashboards Volume / 1RM / Calisthénie": page_dashboards,
    "PR & SAH V2": page_pr_sah,
    "Planning (Annuel / Mésocycles)": page_planning,
    "Synthèse & Recos Globales": page_reco_global,
    "Auto-Séance intelligente": page_auto_seance,
    "Export / Debug": page_export_debug,
}


def main():
    st.set_page_config(page_title="Système Empereur – V3.1", layout="wide")
    st.sidebar.title("Système d'entraînement de l'Empereur – V3.1")
    choix = st.sidebar.radio("Navigation", list(PAGES.keys()))
    st.sidebar.markdown("---")
    st.sidebar.write(f"Modèle : `{TEMPLATE_FILE}`")
    st.sidebar.write(f"Données actives : `{DATA_FILE}`")
    PAGES[choix]()


if __name__ == "__main__":
    main()